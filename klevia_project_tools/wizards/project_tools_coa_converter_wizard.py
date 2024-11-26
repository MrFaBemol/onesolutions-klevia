from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError
from openpyxl import load_workbook

# from xlrd import open_workbook
import base64

import csv
import io
import logging
import pandas as pd

from odoo.tools import file_open

_logger = logging.getLogger(__name__)


class ProjectToolsCoaConverterWizard(models.TransientModel):
    _name = "project.tools.coa.converter.wizard"
    _description = "project.tools.coa.converter.wizard"

    state = fields.Selection(
        selection=[
            ('new', 'New'),
            ('1', '1'),
            ('2', '2'),
            ('3', '3'),
        ],
        default='new',
    )
    file_from = fields.Binary(required=False)
    file_from_name = fields.Char()
    file_to = fields.Binary(required=False)
    file_to_name = fields.Char()
    file_from_id = fields.Many2one(comodel_name='documents.document', required=True)



    def action_read_files(self):
        self.ensure_one()
        # Read from the binary
        try:

            binary_csv = self.file_from  # Replace this with your actual binary CSV content

            # Convert the binary string to a BytesIO object
            binary_stream = io.BytesIO(binary_csv)

            # Read the CSV data using DictReader
            csv_data = []
            with io.TextIOWrapper(binary_stream, encoding='utf-8') as text_stream:
                reader = csv.DictReader(text_stream)
                for row in reader:
                    csv_data.append(row)

            # Example usage of csv_data
            for row in csv_data:
                print(row)



            attachment_id = self.file_from_id.attachment_id

            imp_id = self.env['base_import.import'].create([{
                'res_model': 'account.account',
                'file': self.file_from,
                'file_name': self.file_from_name,
                # 'file_type': file.content_type,
            }])
            imp_id.parse_preview({})
            print(imp_id)
            # imp_id._read_file({})

            # Exemple de binaire (remplacez-le par votre contenu réel)
            binary_excel = self.file_from
            # Charger la chaîne binaire dans un objet BytesIO
            binary_stream = io.BytesIO(binary_excel)
            # Lire le fichier Excel directement avec pandas
            df = pd.read_excel(binary_stream, engine='openpyxl')

            # Afficher le DataFrame
            print("=====================================================")
            print(df)
            print("=====================================================")

            with open(attachment_id._filestore() + '/' + attachment_id.store_fname, 'rb') as f:
                contents = f.read()
                print("=====================================================")
                print(contents)
                print("=====================================================")
                # wb = load_workbook(attachment_id._filestore() + '/' + attachment_id.store_fname)
                try:
                    load_workbook(filename=attachment_id._filestore() + '/' + attachment_id.store_fname)
                except TypeError as e:
                    load_workbook(contents)
                print("=====================================================")
                print(wb)
                print("=====================================================")

            xlsx = io.BytesIO(self.file_from_id)
            wb = load_workbook(xlsx)
            # inputx = io.StringIO()
            # base64.b64decode()
            # inputx.write(base64.decode(self.file_from or b''))
            # book = open_workbook(file_contents=inputx.getvalue())
            raise UserError("%s" % wb)
            # print(wb)
        except TypeError as e:
            raise ValidationError(u'ERROR: {}'.format(e))

        # book = load_workbook(io.BytesIO(self.file_from or b''), data_only=True)
        # print(book)
        print("=====================================================")
        # print(self.file_from)
        # print("=====================================================")
        # print(self.file_to)
        print("=====================================================")

